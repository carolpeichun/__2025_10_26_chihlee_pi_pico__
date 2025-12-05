"""
數據儲存模組
負責將接收到的 MQTT 數據儲存到 Excel 檔案
"""

import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import config


class DataStorage:
    """處理 Excel 數據儲存的類別"""
    
    def __init__(self):
        """初始化數據儲存，確保資料夾和檔案存在"""
        self.data_dir = config.DATA_DIR
        self.excel_path = os.path.join(self.data_dir, config.EXCEL_FILENAME)
        
        # 建立資料夾（如果不存在）
        os.makedirs(self.data_dir, exist_ok=True)
        
        # 初始化或載入 Excel 檔案
        self._init_excel()
    
    def _init_excel(self):
        """初始化 Excel 檔案，如果不存在則建立新檔案"""
        if not os.path.exists(self.excel_path):
            # 建立新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "感測器數據"
            
            # 設定標題列
            headers = ["時間戳記", "電燈狀態", "溫度 (°C)", "溼度 (%)"]
            ws.append(headers)
            
            # 設定標題列樣式
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 調整欄寬
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            wb.save(self.excel_path)
    
    def save_data(self, timestamp, light_status=None, temperature=None, humidity=None):
        """
        儲存數據到 Excel 檔案
        
        Args:
            timestamp: 時間戳記
            light_status: 電燈狀態（開/關）
            temperature: 溫度值
            humidity: 溼度值
        """
        try:
            # 開啟現有的工作簿
            wb = load_workbook(self.excel_path, data_only=False)
            ws = wb.active

            # 檢查是否需要刪除最舊的資料（保持最新的 MAX_RECORDS 筆）
            current_rows = ws.max_row  # 包含標題列
            data_rows = current_rows - 1  # 扣除標題列

            if data_rows >= config.MAX_RECORDS:
                # 需要刪除的行數 = 當前資料行數 - MAX_RECORDS + 1
                rows_to_delete = data_rows - config.MAX_RECORDS + 1
                ws.delete_rows(2, rows_to_delete)
                print(f"已刪除 {rows_to_delete} 筆最舊的資料，保持最新的 {config.MAX_RECORDS} 筆")

            # 準備數據行
            row_data = [
                timestamp,
                light_status if light_status is not None else "",
                temperature if temperature is not None else "",
                humidity if humidity is not None else ""
            ]

            # 新增數據行
            ws.append(row_data)

            # 儲存檔案
            wb.save(self.excel_path)

        except Exception as e:
            print(f"儲存數據時發生錯誤: {e}")
    
    def save_mqtt_message(self, topic, message):
        """
        根據 MQTT 主題儲存對應的數據
        
        Args:
            topic: MQTT 主題
            message: MQTT 訊息內容（JSON 字串或字典）
        """
        try:
            def _to_float(value):
                """嘗試從各種格式中取出數值（例如 '26度' 也能解析為 26）"""
                if value is None:
                    return None
                try:
                    if isinstance(value, (int, float)):
                        return float(value)
                    s = str(value).strip()
                    import re
                    m = re.search(r"-?\d+(?:\.\d+)?", s)
                    if not m:
                        return None
                    return float(m.group(0))
                except Exception:
                    return None

            # 解析訊息
            if isinstance(message, str):
                try:
                    data = json.loads(message)
                except json.JSONDecodeError:
                    # 如果不是 JSON，直接使用原始值
                    data = {"value": message}
            else:
                data = message
            
            # 取得當前時間戳記
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 根據主題名稱判斷並儲存對應數據
            if topic == config.MQTT_TOPICS["light"] or "電燈" in topic:
                # 電燈狀態
                light_status = data.get("status", data.get("value", data.get("light", "")))
                if light_status:
                    self.save_data(timestamp, light_status=str(light_status))
                    print(f"已儲存電燈狀態: {light_status}")
            
            elif topic == config.MQTT_TOPICS["temperature"] or "溫度" in topic:
                # 溫度數據
                temp_value = data.get("temperature", data.get("value", data.get("temp")))
                parsed = _to_float(temp_value)
                if parsed is not None:
                    self.save_data(timestamp, temperature=parsed)
                    print(f"已儲存溫度: {parsed}°C")
            
            elif topic == config.MQTT_TOPICS["humidity"] or "溼度" in topic:
                # 溼度數據
                hum_value = data.get("humidity", data.get("value", data.get("hum")))
                parsed = _to_float(hum_value)
                if parsed is not None:
                    self.save_data(timestamp, humidity=parsed)
                    print(f"已儲存溼度: {parsed}%")
            
            # 如果訊息包含多個數據，一起儲存（優先處理組合訊息）
            if isinstance(data, dict) and ("temperature" in data or "temp" in data or "humidity" in data or "hum" in data or "light" in data or "light_status" in data):
                light_status = data.get("light", data.get("light_status"))
                temperature = data.get("temperature", data.get("temp"))
                humidity = data.get("humidity", data.get("hum"))
                
                # 轉換數值
                temp_parsed = _to_float(temperature) if temperature is not None else None
                hum_parsed = _to_float(humidity) if humidity is not None else None
                
                if any([light_status, temp_parsed is not None, hum_parsed is not None]):
                    self.save_data(timestamp, 
                                 light_status=str(light_status) if light_status else None,
                                 temperature=temp_parsed,
                                 humidity=hum_parsed)
                    print(f"已儲存組合數據: 電燈={light_status}, 溫度={temp_parsed}, 溼度={hum_parsed}")
        
        except Exception as e:
            print(f"處理 MQTT 訊息時發生錯誤: {e}")
    
    def get_history_data(self, limit=100):
        """
        讀取歷史數據
        
        Args:
            limit: 最多讀取的數據筆數（預設100筆）
        
        Returns:
            list: 包含歷史數據的列表，每個元素為字典格式
        """
        try:
            # 顯示實際讀取的檔案路徑（用於調試）
            print(f"嘗試讀取 Excel 檔案: {self.excel_path}")
            print(f"檔案是否存在: {os.path.exists(self.excel_path)}")
            
            if not os.path.exists(self.excel_path):
                print(f"Excel 檔案不存在: {self.excel_path}")
                return []
            
            # 嘗試讀取 Excel 檔案
            # 先嘗試正常模式讀取
            try:
                wb = load_workbook(self.excel_path, read_only=False, data_only=False)
            except Exception as e:
                # 如果遇到 CRC 錯誤，嘗試用 read_only 模式讀取（對損壞檔案更寬容）
                if "CRC" in str(e) or "BadZipFile" in str(e):
                    print(f"警告：Excel 檔案可能有 CRC 錯誤，嘗試用只讀模式讀取: {e}")
                    try:
                        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
                        print("成功用只讀模式讀取檔案")
                    except Exception as e2:
                        print(f"無法讀取 Excel 檔案（即使使用只讀模式）: {e2}")
                        print(f"檔案路徑: {self.excel_path}")
                        print("建議：請檢查檔案是否損壞，或刪除檔案讓系統重新建立")
                        return []
                else:
                    # 其他錯誤直接拋出
                    print(f"讀取 Excel 檔案時發生錯誤: {e}")
                    print(f"檔案路徑: {self.excel_path}")
                    raise
            
            ws = wb.active
            
            # 檢查總行數
            total_rows = ws.max_row
            print(f"Excel 檔案總行數: {total_rows} (包含標題列)")
            
            if total_rows <= 1:
                print("Excel 檔案只有標題列，沒有數據")
                return []
            
            history_data = []
            
            # 從第二行開始讀取（跳過標題列）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) >= 4:
                    timestamp, light_status, temperature, humidity = row[0], row[1], row[2], row[3]
                    
                    # 轉換數據格式，處理各種情況
                    data_point = {
                        'timestamp': str(timestamp) if timestamp else '',
                        'light': None,
                        'temperature': None,
                        'humidity': None
                    }
                    
                    # 處理電燈狀態
                    if light_status:
                        try:
                            data_point['light'] = str(light_status).strip()
                        except:
                            pass
                    
                    # 處理溫度
                    if temperature is not None and temperature != "":
                        try:
                            # 如果是字串，先轉換
                            if isinstance(temperature, str):
                                temp_str = temperature.strip()
                                if temp_str and temp_str != '':
                                    data_point['temperature'] = float(temp_str)
                            else:
                                # 如果是數字，直接轉換
                                data_point['temperature'] = float(temperature)
                        except (ValueError, TypeError) as e:
                            print(f"轉換溫度值失敗: {temperature}, 錯誤: {e}")
                            pass
                    
                    # 處理溼度
                    if humidity is not None and humidity != "":
                        try:
                            # 如果是字串，先轉換
                            if isinstance(humidity, str):
                                hum_str = humidity.strip()
                                if hum_str and hum_str != '':
                                    data_point['humidity'] = float(hum_str)
                            else:
                                # 如果是數字，直接轉換
                                data_point['humidity'] = float(humidity)
                        except (ValueError, TypeError) as e:
                            print(f"轉換溼度值失敗: {humidity}, 錯誤: {e}")
                            pass
                    
                    # 只添加有數據的行
                    if data_point['timestamp'] or data_point['light'] or data_point['temperature'] or data_point['humidity']:
                        history_data.append(data_point)
            
            print(f"成功讀取 {len(history_data)} 筆有效數據")
            
            # 只返回最近的 limit 筆數據
            result = history_data[-limit:] if len(history_data) > limit else history_data
            print(f"返回 {len(result)} 筆數據（限制: {limit}）")
            return result
        
        except Exception as e:
            print(f"讀取歷史數據時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def get_latest_data(self):
        """
        取得最新的數據點（合併最近幾筆數據，確保取得完整的數據）
        
        Returns:
            dict: 包含最新數據的字典
        """
        # 讀取最近 10 筆數據，合併成一個完整的數據點
        history = self.get_history_data(limit=10)
        if not history:
            return {
                'timestamp': '',
                'light': None,
                'temperature': None,
                'humidity': None
            }
        
        # 從最新的數據開始，向後查找，合併所有非 None 的值
        latest = {
            'timestamp': '',
            'light': None,
            'temperature': None,
            'humidity': None
        }
        
        # 從最新到最舊，合併數據
        for item in reversed(history):
            if not latest['timestamp'] and item.get('timestamp'):
                latest['timestamp'] = item.get('timestamp', '')
            if latest['light'] is None and item.get('light') is not None:
                latest['light'] = item.get('light')
            if latest['temperature'] is None and item.get('temperature') is not None:
                latest['temperature'] = item.get('temperature')
            if latest['humidity'] is None and item.get('humidity') is not None:
                latest['humidity'] = item.get('humidity')
            
            # 如果所有值都有了，可以提前結束
            if latest['light'] is not None and latest['temperature'] is not None and latest['humidity'] is not None:
                break
        
        return latest

